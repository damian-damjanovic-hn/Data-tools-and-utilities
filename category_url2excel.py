import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import csv
import html
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

def prettify(text):
    words = text.split("-")
    return ", ".join(word.capitalize() for word in words)

def process_excel_urls(input_file, output_file, pretty_format):
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active

    urls = [row[0].value.strip() for row in ws_in.iter_rows(min_row=1, max_col=1) if row[0].value]

    tree = {}
    for url in urls:
        parts = urlparse(url).path.strip("/").split("/")
        if len(parts) >= 3:
            level1, level2, level3 = parts[-3:]
            if pretty_format:
                level1 = prettify(level1)
                level2 = prettify(level2)
                level3 = prettify(level3)
            tree.setdefault(level1, {}).setdefault(level2, []).append(level3)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Categories"
    ws_out.append(["category_level_1", "category_level_2", "category_level_3"])

    for l1, l2_dict in tree.items():
        for l2, l3_list in l2_dict.items():
            for l3 in l3_list:
                ws_out.append([l1, l2, l3])

    merge_and_center(ws_out)
    wb_out.save(output_file)

def process_csv_categories(input_file, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categories"
    ws.append(["category_level_1", "category_level_2", "category_level_3"])

    with open(input_file, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            level1 = html.unescape(row["category_level_1"]).strip()
            level2 = html.unescape(row["category_level_2"]).strip()
            level3 = html.unescape(row["category_level_3"]).strip()
            ws.append([level1, level2, level3])

    merge_and_center(ws)
    wb.save(output_file)

def merge_and_center(ws):
    for col in range(1, 4):
        start_row = 2
        while start_row <= ws.max_row:
            cell_value = ws.cell(row=start_row, column=col).value
            end_row = start_row
            while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=col).value == cell_value:
                end_row += 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(row=start_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            start_row = end_row + 1

def preview_file(path):
    preview_text.delete("1.0", tk.END)
    try:
        if path.lower().endswith(".csv"):
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    preview_text.insert(tk.END, ", ".join(row) + "\n")
                    if i >= 9:
                        break
        elif path.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                preview_text.insert(tk.END, ", ".join([str(cell) if cell else "" for cell in row]) + "\n")
    except Exception as e:
        preview_text.insert(tk.END, f"Error reading file: {e}")

def run_gui():
    def browse_file():
        path = filedialog.askopenfilename()
        file_path.set(path)
        if path:
            base = os.path.splitext(os.path.basename(path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file.set(f"{base}_converted_{timestamp}.xlsx")
            preview_file(path)

    def generate():
        input_path = file_path.get()
        output_path = output_file.get()
        pretty = pretty_format.get()

        if not input_path or not output_path:
            messagebox.showerror("Error", "Please select input file.")
            return

        try:
            if input_path.lower().endswith(".csv"):
                process_csv_categories(input_path, output_path)
            elif input_path.lower().endswith(".xlsx"):
                process_excel_urls(input_path, output_path, pretty)
            else:
                messagebox.showerror("Error", "Unsupported file format. Please select a .csv or .xlsx file.")
                return
            messagebox.showinfo("Success", f"Excel file '{output_path}' created successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    root = tk.Tk()
    root.title("Category Extractor")

    tk.Label(root, text="Instructions:").grid(row=0, column=0, columnspan=3, sticky="w")
    tk.Label(root, text="1. Select a .csv or .xlsx file containing category data or URLs.").grid(row=1, column=0, columnspan=3, sticky="w")
    tk.Label(root, text="2. The app will auto-detect format and generate output filename.").grid(row=2, column=0, columnspan=3, sticky="w")
    tk.Label(root, text="3. Preview the first few rows and choose formatting option.").grid(row=3, column=0, columnspan=3, sticky="w")

    tk.Label(root, text="Input File:").grid(row=4, column=0, sticky="w")
    file_path = tk.StringVar()
    tk.Entry(root, textvariable=file_path, width=40).grid(row=4, column=1)
    tk.Button(root, text="Browse", command=browse_file).grid(row=4, column=2)

    tk.Label(root, text="Output File:").grid(row=5, column=0, sticky="w")
    output_file = tk.StringVar()
    tk.Entry(root, textvariable=output_file, width=40).grid(row=5, column=1)

    pretty_format = tk.BooleanVar()
    tk.Checkbutton(root, text="Pretty Format (e.g. 'Bathroom, Tiles & Renovations')", variable=pretty_format).grid(row=6, column=0, columnspan=3, sticky="w")

    tk.Button(root, text="Generate Excel", command=generate).grid(row=7, column=1, pady=10)

    tk.Label(root, text="Preview:").grid(row=8, column=0, sticky="nw")
    global preview_text
    preview_text = scrolledtext.ScrolledText(root, width=60, height=10)
    preview_text.grid(row=8, column=1, columnspan=2)

    root.mainloop()

run_gui()
